"""
OrgImporter — 从 AnyShare 导出的用户 Excel 批量创建 BISHENG 组织架构

支持两种数据来源：
  1. Excel 文件（anyshare 导出格式）
  2. AnyShare API 实时拉取

用法（通过 run.py）：
  python run.py --import-org                          # 使用 config 里配置的 excel_path
  python run.py --import-org path/to/users.xlsx       # 指定 Excel 文件
"""

from __future__ import annotations

import base64
import hashlib
import hmac
import json
import logging
import re
import time
from pathlib import Path
from typing import TYPE_CHECKING

import httpx

logger = logging.getLogger(__name__)

if TYPE_CHECKING:
    from app.config import AppConfig


class OrgImporter:
    """从 Excel 导出文件批量创建 BISHENG 部门和用户。"""

    ROOT_DEPT_INT_ID = 1       # BISHENG 系统根部门整数 id
    DEFAULT_ROLE_IDS = [2]     # 普通用户角色
    DEFAULT_PASSWORD  = "Sync@123456"

    def __init__(self, config: "AppConfig"):
        self._cfg    = config
        self._base   = config.bisheng.base_url.rstrip("/")
        self._secret = getattr(config.bisheng, "jwt_secret", "")
        self._timeout = getattr(config.bisheng, "timeout", 30)

        # 运行时状态
        self._rsa_pub    = None          # cryptography RSA public key object
        self._pwd_enc    = None          # 加密后的密码（复用）
        self._dept_map: dict[str, int]   = {}   # path -> int id
        self._dept_id_map: dict[str, str] = {}  # path -> BS@xxx string
        self._seen_dept_ids: set[int]    = set()

    # ── Token ─────────────────────────────────────────────────────

    def _make_token(self) -> str:
        """生成 BISHENG JWT（管理员）。"""
        cfg = self._cfg.bisheng

        def b64url(d: bytes | str) -> str:
            if isinstance(d, str): d = d.encode()
            return base64.urlsafe_b64encode(d).rstrip(b"=").decode()

        sub = json.dumps({
            "user_id":       getattr(cfg, "jwt_admin_user_id",       1),
            "user_name":     getattr(cfg, "jwt_admin_user_name",      "admin"),
            "tenant_id":     getattr(cfg, "jwt_admin_tenant_id",      1),
            "token_version": getattr(cfg, "jwt_admin_token_version",  1),
        })
        expire = int(time.time()) + getattr(cfg, "jwt_expire_seconds", 86400)
        h = b64url(json.dumps({"alg": "HS256", "typ": "JWT"}, separators=(",", ":")))
        p = b64url(json.dumps({"sub": sub, "exp": expire, "iss": getattr(cfg, "jwt_issuer", "bisheng")},
                              separators=(",", ":")))
        sig = hmac.new(self._secret.encode(), f"{h}.{p}".encode(), hashlib.sha256).digest()
        return f"{h}.{p}.{b64url(sig)}"

    def _cookies(self) -> dict:
        return {"access_token_cookie": self._make_token()}

    # ── RSA 密码加密 ──────────────────────────────────────────────

    def _init_rsa(self) -> None:
        """从 BISHENG 获取 RSA 公钥并加密默认密码。"""
        from cryptography.hazmat.primitives.asymmetric import padding as rsa_padding
        from cryptography.hazmat.primitives.asymmetric.rsa import RSAPublicNumbers
        from cryptography.hazmat.backends import default_backend

        r = httpx.get(f"{self._base}/api/v1/user/public_key",
                      cookies=self._cookies(), timeout=self._timeout)
        pem = r.json()["data"]["public_key"]
        der = base64.b64decode(re.sub(r"-----.*?-----|\s", "", pem))

        def _parse_pkcs1(der: bytes):
            pos = 0
            def read_tlv():
                nonlocal pos
                t = der[pos]; pos += 1
                l = der[pos]; pos += 1
                if l & 0x80:
                    nb = l & 0x7f
                    l = int.from_bytes(der[pos:pos+nb], "big"); pos += nb
                v = der[pos:pos+l]; pos += l
                return t, v
            _, seq = read_tlv()
            def read_int(d):
                p = 0; t = d[p]; p += 1; l = d[p]; p += 1
                if l & 0x80:
                    nb = l & 0x7f
                    l = int.from_bytes(d[p:p+nb], "big"); p += nb
                v = d[p:p+l]; p += l
                return int.from_bytes(v, "big"), d[p:]
            n, rest = read_int(seq)
            e, _    = read_int(rest)
            return n, e

        n, e = _parse_pkcs1(der)
        self._rsa_pub = RSAPublicNumbers(e, n).public_key(default_backend())
        encrypted = self._rsa_pub.encrypt(self.DEFAULT_PASSWORD.encode(), rsa_padding.PKCS1v15())
        self._pwd_enc = base64.b64encode(encrypted).decode()
        logger.debug("RSA public key loaded and password encrypted")

    # ── 读取 Excel ────────────────────────────────────────────────

    @staticmethod
    def read_excel(excel_path: str | Path) -> list[dict]:
        """读取 AnyShare 导出的用户 Excel，返回用户列表。

        每条记录：{username, display, dept}
          username = AnyShare 用户名（英文 ID）→ BISHENG person_id
          display  = AnyShare 显示名（中文名）→ BISHENG user_name
          dept     = 完整部门路径，用 / 分隔
        """
        from openpyxl import load_workbook

        wb = load_workbook(str(excel_path), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # 找表头行
        data_start = 0
        for i, row in enumerate(rows):
            if row[0] == "用户名":
                data_start = i + 1
                break

        users = []
        for row in rows[data_start:]:
            if not row[0]:
                continue
            username = str(row[0]).strip()
            display  = str(row[1]).strip() if row[1] else username
            dept     = str(row[2]).strip() if row[2] else ""
            users.append({"username": username, "display": display, "dept": dept})

        logger.info(f"Excel loaded: {len(users)} users from {excel_path}")
        return users

    # ── 部门操作 ──────────────────────────────────────────────────

    def _load_dept_children(self, parent_int_id: int, parent_path: str = "") -> None:
        """递归加载部门子树，填充 dept_map / dept_id_map。"""
        r = httpx.get(
            f"{self._base}/api/v1/departments/children",
            params={"parent_id": parent_int_id, "include_archived": "true"},
            cookies=self._cookies(), timeout=self._timeout,
        )
        nodes = r.json().get("data") or []
        if not isinstance(nodes, list):
            return
        for node in nodes:
            nid  = node.get("id")
            bs   = node.get("dept_id", "")
            nm   = node.get("name", "")
            path = (parent_path + "/" + nm) if parent_path else nm
            if nid not in self._seen_dept_ids:
                self._seen_dept_ids.add(nid)
                self._dept_map[path]    = nid
                self._dept_id_map[path] = bs
            if node.get("has_children"):
                self._load_dept_children(nid, path)

    def _ensure_dept(self, path: str) -> str | None:
        """确保部门路径存在，返回 BS@xxx 字符串 id。按需创建。"""
        if path in self._dept_id_map:
            return self._dept_id_map[path]

        parts     = path.split("/")
        name      = parts[-1]
        par_path  = "/".join(parts[:-1]) if len(parts) > 1 else ""
        parent_id = self._dept_map.get(par_path, self.ROOT_DEPT_INT_ID) if par_path else self.ROOT_DEPT_INT_ID

        r = httpx.post(
            f"{self._base}/api/v1/departments/",
            json={"name": name, "parent_id": parent_id},
            cookies=self._cookies(), timeout=self._timeout,
        )
        resp = r.json()
        if resp.get("status_code") == 200:
            data = resp["data"]
            nid  = data["id"]
            bs   = data.get("dept_id", "")
            self._dept_map[path]    = nid
            self._dept_id_map[path] = bs
            self._seen_dept_ids.add(nid)
            return bs
        else:
            msg = str(resp.get("status_message", ""))
            if "exist" in msg.lower() or "already" in msg.lower():
                # 已存在，重新加载该层找到 id
                self._load_dept_children(parent_id, par_path)
                return self._dept_id_map.get(path)
            logger.warning(f"Failed to create dept [{path}]: {msg}")
            return None

    def _build_dept_tree(self, users: list[dict]) -> None:
        """从用户列表提取所有部门路径并按层级创建。"""
        dept_paths: set[str] = set()
        for u in users:
            dept = u.get("dept", "").split(",")[0].strip()
            parts = [p.strip() for p in dept.split("/") if p.strip()]
            if len(parts) <= 1:
                continue
            parts = parts[1:]   # 跳过根公司名（BISHENG 根节点已存在）
            path = ""
            for p in parts:
                path = path + "/" + p if path else p
                dept_paths.add(path)

        sorted_paths = sorted(dept_paths, key=lambda x: x.count("/"))
        created = skipped = failed = 0

        for path in sorted_paths:
            if path in self._dept_map:
                skipped += 1
                continue
            bs = self._ensure_dept(path)
            if bs:
                created += 1
                if created % 200 == 0:
                    logger.info(f"  Departments created: {created}...")
            else:
                failed += 1

            time.sleep(0.03)

        logger.info(f"Departments: created={created} skipped={skipped} failed={failed}")
        return created, skipped, failed

    # ── 用户操作 ──────────────────────────────────────────────────

    def _get_existing_users(self) -> set[str]:
        """返回 BISHENG 中已有的 person_id (external_id) 集合。"""
        existing: set[str] = set()
        page = 1
        while True:
            r = httpx.get(
                f"{self._base}/api/v1/user/list",
                params={"page": page, "page_size": 500},
                cookies=self._cookies(), timeout=self._timeout,
            )
            data  = r.json().get("data", {})
            items = data.get("data", [])
            for u in items:
                pid = u.get("external_id") or u.get("person_id") or ""
                if pid:
                    existing.add(pid)
            if page * 500 >= data.get("total", 0):
                break
            page += 1
        return existing

    def _create_user(self, u: dict) -> bool:
        """创建单个用户，返回是否成功。"""
        dept = u.get("dept", "").split(",")[0].strip()
        parts = [p.strip() for p in dept.split("/") if p.strip()]
        bs_dept = "BS@root"
        if len(parts) > 1:
            sub_path = "/".join(parts[1:])
            bs_dept  = self._dept_id_map.get(sub_path, "BS@root")

        body = {
            "dept_id":   bs_dept,
            "user_name": u["display"],    # 中文显示名
            "person_id": u["username"],   # 英文 ID
            "password":  self._pwd_enc,
            "role_ids":  self.DEFAULT_ROLE_IDS,
        }
        r = httpx.post(
            f"{self._base}/api/v1/departments/local-members",
            json=body, cookies=self._cookies(), timeout=self._timeout,
        )
        sc = r.json().get("status_code")
        return sc == 200

    # ── 主入口 ────────────────────────────────────────────────────

    def run(self, excel_path: str | Path) -> dict:
        """执行完整的组织架构导入流程。

        Args:
            excel_path: AnyShare 导出的用户 Excel 文件路径

        Returns:
            {dept_created, dept_skipped, dept_failed,
             user_ok, user_skipped, user_failed}
        """
        logger.info(f"=== OrgImporter start: {excel_path} ===")

        # 1. RSA 公钥 + 密码加密
        logger.info("Fetching RSA public key...")
        self._init_rsa()

        # 2. 读取 Excel
        logger.info("Reading Excel...")
        users = self.read_excel(excel_path)
        logger.info(f"  {len(users)} users")

        # 3. 加载已有部门
        logger.info("Loading existing departments...")
        self._load_dept_children(self.ROOT_DEPT_INT_ID)
        logger.info(f"  {len(self._dept_map)} existing departments")

        # 4. 创建缺失部门
        logger.info("Building department tree...")
        created_d, skipped_d, failed_d = self._build_dept_tree(users)

        # 5. 重新加载部门映射（确保 dept_id_map 完整）
        self._dept_map.clear()
        self._dept_id_map.clear()
        self._seen_dept_ids.clear()
        self._load_dept_children(self.ROOT_DEPT_INT_ID)
        logger.info(f"Dept map reloaded: {len(self._dept_map)} departments")

        # 6. 查已有用户
        logger.info("Checking existing users...")
        existing = self._get_existing_users()
        logger.info(f"  {len(existing)} existing users")

        to_create = [u for u in users if u["username"] not in existing]
        logger.info(f"  {len(to_create)} to create, {len(users)-len(to_create)} skip")

        # 7. 批量创建用户
        logger.info("Creating users...")
        ok = ng = skip_dup = 0
        for i, u in enumerate(to_create):
            if i % 500 == 0 and i > 0:
                logger.info(f"  [{i}/{len(to_create)}] ok={ok} dup={skip_dup} fail={ng}")
            try:
                if self._create_user(u):
                    ok += 1
                else:
                    skip_dup += 1
            except Exception as e:
                ng += 1
                logger.warning(f"  User {u['username']} failed: {e}")

            if i % 20 == 19:
                time.sleep(0.1)

        result = {
            "dept_created": created_d, "dept_skipped": skipped_d, "dept_failed": failed_d,
            "user_ok": ok, "user_skipped": len(users) - len(to_create) + skip_dup,
            "user_failed": ng,
        }
        logger.info(f"=== OrgImporter done: {result} ===")
        return result
